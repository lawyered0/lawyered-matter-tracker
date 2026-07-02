#!/usr/bin/env python3
"""tracker_write.py — deterministic write-guard CLI for the matter tracker.

Every mutation of matter-tracker.xlsx made by Claude skills MUST go through
this script instead of ad-hoc openpyxl code. All safety properties live in
this code path:

  * refuses to write while Excel has the file open (owner lock file),
    unless --force-unlocked
  * timestamped backup into backups/ beside the tracker before every write
    (pruned by mtime, most recent 200 kept)
  * atomic save (temp file + os.replace) so a crash never leaves a
    half-written tracker
  * post-write integrity check via validate_tracker.py; a FAIL prints the
    validator output and exits 3 with the backup path for easy restore
  * column format contract enforced on the values being written (exit 2)
  * Last Activity only ever moves forward, and never beyond today

Usage:
  tracker_write.py update      --tracker PATH --file-no N --set "COLUMN=value" [--set ...]
  tracker_write.py timeline    --tracker PATH --file-no N --date YYYY-MM-DD --text "..."
  tracker_write.py new-matter  --tracker PATH --client "..." --description "..."
                               [--opposing ...] [--email ...] [--phone ...]
                               [--matter-type ...] [--matter-folder ...] [--date-opened YYYY-MM-DD]
  tracker_write.py close       --tracker PATH --file-no N
  tracker_write.py reopen      --tracker PATH --file-no N
  tracker_write.py court-deadline add     --tracker PATH --file-no N
                               (--date YYYY-MM-DD | --anchor "trial date" --offset-days -30)
                               --description "..." [--source "..."]
  tracker_write.py court-deadline remove  --tracker PATH --file-no N --index I
  tracker_write.py court-deadline resolve --tracker PATH --file-no N --index I --date YYYY-MM-DD

Global flags (valid on every subcommand): --tracker PATH (required),
--dry-run (report what would change, write nothing), --json (machine-readable
result on stdout), --force-unlocked (override the Excel lock check).

Exit codes: 0 = success, 1 = operational error (file/matter not found, lock,
structure), 2 = value rejected by the column format contract (or argparse
usage error), 3 = post-write validation FAILED (restore from the backup path
printed in the message).
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Schema contract (must match validate_tracker.py / the live webapp)
# ---------------------------------------------------------------------------

EXPECTED_HEADERS = [
    "File #", "Client Name", "Matter Description", "Status",
    "Date Opened", "Date Closed", "Last Activity", "Opposing Party",
    "Next Action / Deadline", "Timeline", "Client ID Verified",
    "Conflict Check Done", "Client Email", "Client Phone",
    "Client Address", "Discovery Date", "Limitation Statute",
    "Limitation Deadline", "Court Deadlines", "Matter Folder",
    "Other Parties / Related Persons", "Matter Type",
]
# Extension columns (created on first need by the skills; writable via update
# once the header exists on the sheet):
EXTENSION_COLUMNS = ["Related Matters", "Clio Synced"]
DATE_COLUMNS = {"Date Opened", "Date Closed", "Last Activity",
                "Discovery Date", "Limitation Deadline"}
SHEETS = ["Open Matters", "Closed Matters"]
MAX_BACKUPS = 200
NEXT_ACTION_HARD_CAP = 200
NEXT_ACTION_WARN = 80

EXIT_OK = 0
EXIT_ERROR = 1       # operational: file/matter not found, lock, bad structure
EXIT_CONTRACT = 2    # value rejected by the column format contract
EXIT_VALIDATOR = 3   # post-write validate_tracker.py FAIL

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
VALIDATOR_PATH = os.path.join(SCRIPT_DIR, "validate_tracker.py")


class GuardError(Exception):
    """Fatal error; .code is the process exit code."""

    def __init__(self, message, code=EXIT_ERROR):
        super().__init__(message)
        self.code = code


def contract_error(msg):
    return GuardError(msg, EXIT_CONTRACT)


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def today_iso():
    return date.today().strftime("%Y-%m-%d")


def to_iso(val):
    """Normalize a cell value to a string; dates become YYYY-MM-DD."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def is_valid_iso_date(s):
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", s or ""):
        return False
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def require_iso_date(s, what):
    """Contract check: s must be a real YYYY-MM-DD date."""
    if not is_valid_iso_date(s):
        raise contract_error(
            f"{what} must be a valid date in YYYY-MM-DD format, got {s!r}")
    return s


def validate_date_column_value(column, value):
    """Contract for the five date columns: YYYY-MM-DD or empty."""
    if value == "":
        return value
    require_iso_date(value, f'"{column}"')
    if column == "Last Activity" and value > today_iso():
        raise contract_error(
            f'"Last Activity" may never be a future date (got {value}, today is {today_iso()})')
    return value


def validate_next_action(value):
    """Contract for "Next Action / Deadline". Returns list of warnings."""
    warnings = []
    if value == "":
        return warnings
    if "\n" in value or "\r" in value:
        raise contract_error(
            '"Next Action / Deadline" must be a single line (newline found). '
            "Put history in the Timeline instead.")
    if len(value) > NEXT_ACTION_HARD_CAP:
        raise contract_error(
            f'"Next Action / Deadline" is {len(value)} chars; hard cap is '
            f"{NEXT_ACTION_HARD_CAP}. Trim it — details belong in the Timeline.")
    if len(value) > NEXT_ACTION_WARN:
        warnings.append(
            f'"Next Action / Deadline" is {len(value)} chars (> {NEXT_ACTION_WARN}); '
            "consider shortening.")
    m = re.match(r"^\s*(\d{4}-\d{2}-\d{2})", value)
    if m:
        require_iso_date(m.group(1), 'leading date in "Next Action / Deadline"')
        if not re.match(r"^\d{4}-\d{2}-\d{2}: \S", value):
            raise contract_error(
                '"Next Action / Deadline" starting with a date must use the '
                'exact form "YYYY-MM-DD: <action>" (date, colon, single space), '
                f"got {value!r}")
    return warnings


# ---------------------------------------------------------------------------
# Workbook plumbing
# ---------------------------------------------------------------------------

def check_lock(tracker, force_unlocked):
    """Refuse to write while Excel holds an owner lock file, unless forced."""
    lock = os.path.join(os.path.dirname(tracker), "~$" + os.path.basename(tracker))
    if os.path.exists(lock) and not force_unlocked:
        raise GuardError(
            f"Tracker appears to be open in Excel (owner lock file exists: {lock}). "
            "Close it in Excel first, or re-run with --force-unlocked if you are "
            "certain the lock is stale.", EXIT_ERROR)


def open_tracker(tracker):
    if not os.path.isfile(tracker):
        raise GuardError(f"Tracker file not found: {tracker}", EXIT_ERROR)
    try:
        wb = load_workbook(tracker)
    except Exception as e:
        raise GuardError(f"Cannot open tracker {tracker}: {e}", EXIT_ERROR)
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            raise GuardError(
                f"Tracker is missing required sheet '{sheet}' "
                f"(found: {wb.sheetnames})", EXIT_ERROR)
    return wb


def sheet_headers(ws):
    return [c.value for c in ws[1]]


def col_index(ws, headers, column):
    """1-based column index for a header; error if the sheet lacks it."""
    if column not in headers:
        raise GuardError(
            f"Sheet '{ws.title}' has no column '{column}' — tracker structure "
            "is out of date. Fix the headers before writing.", EXIT_ERROR)
    return headers.index(column) + 1


def find_matter(wb, file_no):
    """Locate a File # across both sheets. Errors if on neither or both."""
    hits = []
    for sheet in SHEETS:
        ws = wb[sheet]
        headers = sheet_headers(ws)
        fc = col_index(ws, headers, "File #")
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=fc).value
            if v is not None and str(v).strip() == str(file_no).strip():
                hits.append((sheet, r))
    if not hits:
        raise GuardError(
            f"Matter with File # {file_no} not found on Open Matters or "
            "Closed Matters.", EXIT_ERROR)
    if len(hits) > 1:
        where = ", ".join(f"{s} row {r}" for s, r in hits)
        raise GuardError(
            f"File # {file_no} appears more than once ({where}) — tracker "
            "integrity problem; fix duplicates before writing.", EXIT_ERROR)
    return hits[0]


def backup_spreadsheet(tracker):
    """Timestamped copy into backups/ beside the tracker; prune oldest by
    mtime keeping MAX_BACKUPS."""
    backup_dir = os.path.join(os.path.dirname(tracker), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    stem = os.path.splitext(os.path.basename(tracker))[0]
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"{stem}_{ts}.xlsx")
    n = 1
    while os.path.exists(backup_path):  # several writes in the same second
        backup_path = os.path.join(backup_dir, f"{stem}_{ts}_{n}.xlsx")
        n += 1
    shutil.copy2(tracker, backup_path)
    # Prune by actual mtime, not filename, so retention works regardless of
    # naming convention (other tools drop backups here under other names).
    try:
        backups = sorted(
            (os.path.join(backup_dir, f) for f in os.listdir(backup_dir)
             if f.endswith(".xlsx")),
            key=os.path.getmtime, reverse=True)
        for old in backups[MAX_BACKUPS:]:
            try:
                os.remove(old)
            except OSError:
                pass  # already gone or locked; not worth failing a save over
    except OSError:
        pass  # never let backup cleanup break the write it's protecting
    return backup_path


def atomic_save(wb, tracker):
    """Temp file + os.replace so readers never see a half-written tracker."""
    tmp_path = f"{tracker}.{os.getpid()}.tmp"
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, tracker)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def run_validator(tracker, backup_path):
    """Run validate_tracker.py after the write; exit 3 loudly on FAIL."""
    if not os.path.isfile(VALIDATOR_PATH):
        raise GuardError(
            f"validate_tracker.py not found at {VALIDATOR_PATH} — cannot "
            f"verify the write. Backup made before this write: {backup_path}",
            EXIT_VALIDATOR)
    proc = subprocess.run(
        [sys.executable, VALIDATOR_PATH, tracker, backup_path],
        capture_output=True, text=True)
    if proc.returncode != 0:
        sys.stderr.write(proc.stdout)
        sys.stderr.write(proc.stderr)
        raise GuardError(
            "POST-WRITE VALIDATION FAILED — the tracker may now be inconsistent. "
            f"Restore from the backup made just before this write: {backup_path}",
            EXIT_VALIDATOR)
    return "PASS"


# ---------------------------------------------------------------------------
# Shared mutation helpers
# ---------------------------------------------------------------------------

def append_timeline(ws, row, headers, entry_date, text):
    """Append "\\ndate: text" to the Timeline cell (no leading newline when
    the cell was empty). Returns the appended line."""
    tc = col_index(ws, headers, "Timeline")
    current = to_iso(ws.cell(row=row, column=tc).value)
    line = f"{entry_date}: {text}"
    ws.cell(row=row, column=tc).value = (current + "\n" + line) if current else line
    return line


def bump_last_activity(ws, row, headers, event_date):
    """Last Activity = max(current, event_date), capped at today.
    Returns the value now stored."""
    lc = col_index(ws, headers, "Last Activity")
    event_date = min(event_date, today_iso())
    current = to_iso(ws.cell(row=row, column=lc).value)
    if is_valid_iso_date(current):
        new = max(current, event_date)
    else:
        new = event_date  # blank or unparseable — replace with the event date
    ws.cell(row=row, column=lc).value = new
    return new


def parse_court_deadlines(raw):
    """Col S contract: empty/blank -> []; otherwise a JSON array of objects."""
    s = to_iso(raw)
    if not s:
        return []
    try:
        arr = json.loads(s)
    except ValueError as e:
        raise contract_error(
            f'"Court Deadlines" cell does not contain valid JSON ({e}). '
            "Refusing to overwrite it — fix the cell first.")
    if not isinstance(arr, list):
        raise contract_error(
            '"Court Deadlines" cell must be a JSON array; refusing to overwrite it.')
    return arr


def sort_court_deadlines(arr):
    """Dated entries sorted by date; anchored entries (no date key) last,
    keeping their relative order."""
    dated = sorted((e for e in arr if isinstance(e, dict) and e.get("date")),
                   key=lambda e: e["date"])
    anchored = [e for e in arr if not (isinstance(e, dict) and e.get("date"))]
    return dated + anchored


def write_court_deadlines(ws, row, headers, arr):
    sc = col_index(ws, headers, "Court Deadlines")
    ws.cell(row=row, column=sc).value = (
        json.dumps(arr, ensure_ascii=False) if arr else None)


def read_cell(ws, row, headers, column):
    return to_iso(ws.cell(row=row, column=col_index(ws, headers, column)).value)


def set_cell(ws, row, headers, column, value):
    ws.cell(row=row, column=col_index(ws, headers, column)).value = (
        value if value != "" else None)


# ---------------------------------------------------------------------------
# Subcommand implementations
# Each returns (file_no, changes_dict, warnings_list, summary_str) after
# mutating the in-memory workbook. The common driver handles dry-run/backup/
# save/validate.
# ---------------------------------------------------------------------------

def cmd_update(args, wb):
    sheet, row = find_matter(wb, args.file_no)
    ws = wb[sheet]
    headers = sheet_headers(ws)
    warnings = []
    pending = {}

    for spec in args.sets:
        if "=" not in spec:
            raise contract_error(
                f'--set expects "COLUMN=value", got {spec!r}')
        column, value = spec.split("=", 1)
        column = column.strip()

        if column not in EXPECTED_HEADERS and column not in EXTENSION_COLUMNS:
            raise contract_error(
                f"Unknown column {column!r}. Valid columns: "
                + ", ".join(f'"{h}"' for h in EXPECTED_HEADERS + EXTENSION_COLUMNS))
        if column == "File #":
            raise contract_error(
                '"File #" is immutable — it can never be changed via update.')
        if column == "Status":
            raise contract_error(
                '"Status" cannot be edited directly. Use the close/reopen '
                "subcommands, which move the row between sheets and keep "
                "Status, Date Closed and the Timeline consistent.")
        if column == "Court Deadlines":
            raise contract_error(
                '"Court Deadlines" cannot be edited via update. Use the '
                "court-deadline add/remove/resolve subcommands, which keep "
                "the JSON structure and sort order intact.")
        if column in DATE_COLUMNS:
            value = validate_date_column_value(column, value.strip())
        elif column == "Next Action / Deadline":
            warnings.extend(validate_next_action(value))
        pending[column] = value

    changes = {}
    for column, value in pending.items():
        old = read_cell(ws, row, headers, column)
        set_cell(ws, row, headers, column, value)
        changes[column] = {"old": old, "new": value}
    summary = "; ".join(f"{c}={v['new']!r}" for c, v in changes.items())
    return args.file_no, changes, warnings, f"[{sheet}] set {summary}"


def cmd_timeline(args, wb):
    entry_date = require_iso_date(args.date, "--date")
    if entry_date > today_iso():
        raise contract_error(
            f"Timeline entries record what has happened — --date {entry_date} "
            f"is in the future (today is {today_iso()}). Future obligations "
            "belong in Next Action / Deadline or Court Deadlines.")
    text = args.text.strip()
    if not text:
        raise contract_error("--text must not be empty")
    if "\n" in text or "\r" in text:
        raise contract_error(
            "Timeline entry text must be a single line — each entry is one "
            '"date: text" line in the Timeline cell.')

    sheet, row = find_matter(wb, args.file_no)
    ws = wb[sheet]
    headers = sheet_headers(ws)
    old_la = read_cell(ws, row, headers, "Last Activity")
    line = append_timeline(ws, row, headers, entry_date, text)
    new_la = bump_last_activity(ws, row, headers, entry_date)
    changes = {
        "Timeline": {"appended": line},
        "Last Activity": {"old": old_la, "new": new_la},
    }
    return args.file_no, changes, [], f"[{sheet}] timeline += {line!r}; Last Activity={new_la}"


def generate_next_file_no(wb):
    """Next File # = current year + (max NNN across BOTH sheets for that
    year) + 1, zero-padded to 3."""
    year = str(date.today().year)
    max_seq = 0
    for sheet in SHEETS:
        ws = wb[sheet]
        headers = sheet_headers(ws)
        fc = col_index(ws, headers, "File #")
        for r in range(2, ws.max_row + 1):
            fn = to_iso(ws.cell(row=r, column=fc).value)
            m = re.match(rf"^{year}-(\d+)$", fn)
            if m:
                max_seq = max(max_seq, int(m.group(1)))
    return f"{year}-{max_seq + 1:03d}"


def cmd_new_matter(args, wb):
    client = args.client.strip()
    description = args.description.strip()
    if not client:
        raise contract_error("--client must not be empty")
    if not description:
        raise contract_error("--description must not be empty")
    date_opened = (args.date_opened or today_iso()).strip()
    require_iso_date(date_opened, "--date-opened")

    file_no = generate_next_file_no(wb)
    # Belt and braces: verify uniqueness across BOTH sheets before writing.
    for sheet in SHEETS:
        ws = wb[sheet]
        headers = sheet_headers(ws)
        fc = col_index(ws, headers, "File #")
        for r in range(2, ws.max_row + 1):
            if to_iso(ws.cell(row=r, column=fc).value) == file_no:
                raise GuardError(
                    f"Generated File # {file_no} already exists on {sheet} "
                    f"row {r} — refusing to create a duplicate.", EXIT_ERROR)

    ws = wb["Open Matters"]
    headers = sheet_headers(ws)
    values = {
        "File #": file_no,
        "Client Name": client,
        "Matter Description": description,
        "Status": "Open",
        "Date Opened": date_opened,
        "Conflict Check Done": "✓",
        "Client ID Verified": "Pending",
        "Opposing Party": (args.opposing or "").strip(),
        "Client Email": (args.email or "").strip(),
        "Client Phone": (args.phone or "").strip(),
        "Matter Type": (args.matter_type or "").strip(),
        "Matter Folder": (args.matter_folder or "").strip(),
    }
    new_row = ws.max_row + 1
    for column, value in values.items():
        if value:
            set_cell(ws, new_row, headers, column, value)
    changes = {c: {"old": "", "new": v} for c, v in values.items() if v}
    return file_no, changes, [], (
        f"created {file_no} on Open Matters row {new_row}: {client} — {description}")


def _move_matter(args, wb, from_sheet, to_sheet, updates, timeline_text):
    sheet, row = find_matter(wb, args.file_no)
    if sheet != from_sheet:
        verb = "closed" if to_sheet == "Closed Matters" else "reopened"
        raise GuardError(
            f"Matter {args.file_no} is on '{sheet}', not '{from_sheet}' — "
            f"it cannot be {verb} from there (is it already {verb}?).", EXIT_ERROR)
    src = wb[from_sheet]
    dst = wb[to_sheet]
    src_headers = sheet_headers(src)
    dst_headers = sheet_headers(dst)
    col_index(dst, dst_headers, "File #")  # structural sanity on destination

    row_data = {}
    for i, h in enumerate(src_headers):
        if h is not None:
            row_data[h] = src.cell(row=row, column=i + 1).value
    for k, v in updates.items():
        row_data[k] = v

    # A move must never silently drop data: any source column carrying a
    # value must exist on the destination. Known extension columns get their
    # header auto-created; anything else is an error.
    for column, value in row_data.items():
        if column not in dst_headers and value not in (None, ""):
            if column in EXTENSION_COLUMNS:
                new_col = len([h for h in dst_headers if h is not None]) + 1
                dst.cell(row=1, column=new_col).value = column
                dst_headers = sheet_headers(dst)
            else:
                raise GuardError(
                    f"Sheet '{to_sheet}' has no column {column!r} but the row "
                    f"carries a value there — the move would silently drop it. "
                    f"Add the header to '{to_sheet}' first.", EXIT_ERROR)

    dst_row = dst.max_row + 1
    for column, value in row_data.items():
        if column in dst_headers:
            dst.cell(row=dst_row, column=dst_headers.index(column) + 1).value = value
    src.delete_rows(row)

    today = today_iso()
    line = append_timeline(dst, dst_row, dst_headers, today, timeline_text)
    new_la = bump_last_activity(dst, dst_row, dst_headers, today)

    changes = {"Sheet": {"old": from_sheet, "new": to_sheet}}
    for k, v in updates.items():
        changes[k] = {"new": to_iso(v)}
    changes["Timeline"] = {"appended": line}
    changes["Last Activity"] = {"new": new_la}
    summary = (f"moved {from_sheet} -> {to_sheet}; "
               + "; ".join(f"{k}={to_iso(v)!r}" for k, v in updates.items())
               + f"; timeline += {line!r}; Last Activity={new_la}")
    return args.file_no, changes, [], summary


def cmd_close(args, wb):
    return _move_matter(args, wb, "Open Matters", "Closed Matters",
                        {"Status": "Closed", "Date Closed": today_iso()},
                        "Matter closed")


def cmd_reopen(args, wb):
    return _move_matter(args, wb, "Closed Matters", "Open Matters",
                        {"Status": "Open", "Date Closed": None},
                        "Matter reopened")


def _load_deadlines(args, wb):
    sheet, row = find_matter(wb, args.file_no)
    ws = wb[sheet]
    headers = sheet_headers(ws)
    sc = col_index(ws, headers, "Court Deadlines")
    arr = parse_court_deadlines(ws.cell(row=row, column=sc).value)
    return sheet, row, ws, headers, arr


def _finish_deadlines(args, ws, row, headers, arr, sheet, action_desc):
    arr = sort_court_deadlines(arr)
    write_court_deadlines(ws, row, headers, arr)
    new_la = bump_last_activity(ws, row, headers, today_iso())
    changes = {
        "Court Deadlines": {"entries": arr, "count": len(arr)},
        "Last Activity": {"new": new_la},
    }
    summary = (f"[{sheet}] {action_desc}; Court Deadlines now has "
               f"{len(arr)} entr{'y' if len(arr) == 1 else 'ies'}; "
               f"Last Activity={new_la}")
    return args.file_no, changes, [], summary


def cmd_cd_add(args, wb):
    description = (args.description or "").strip()
    if not description:
        raise contract_error("--description must not be empty")
    source = (args.source or "").strip()

    anchored_bits = args.anchor is not None or args.offset_days is not None
    if args.date and anchored_bits:
        raise contract_error(
            "Give either --date (a known date) or --anchor + --offset-days "
            "(an anchored deadline), not both.")
    if args.date:
        entry_date = require_iso_date(args.date.strip(), "--date")
        entry = {"date": entry_date, "description": description, "source": source}
        desc = f"added dated deadline {entry_date}: {description!r}"
    elif args.anchor is not None or args.offset_days is not None:
        anchor = (args.anchor or "").strip()
        if not anchor:
            raise contract_error("--anchor must not be empty")
        if args.offset_days is None:
            raise contract_error("--anchor requires --offset-days (integer, may be negative)")
        entry = {"anchor": anchor, "offset_days": args.offset_days,
                 "description": description, "source": source}
        desc = (f"added anchored deadline ({anchor!r} {args.offset_days:+d} days): "
                f"{description!r}")
    else:
        raise contract_error(
            "court-deadline add needs --date YYYY-MM-DD, or --anchor plus --offset-days.")

    sheet, row, ws, headers, arr = _load_deadlines(args, wb)
    arr.append(entry)
    return _finish_deadlines(args, ws, row, headers, arr, sheet, desc)


def _check_index(arr, index, file_no):
    if not arr:
        raise contract_error(
            f"Matter {file_no} has no court deadlines — nothing at index {index}.")
    if not (0 <= index < len(arr)):
        raise contract_error(
            f"--index {index} is out of range; valid indexes are 0..{len(arr) - 1} "
            f"({len(arr)} entr{'y' if len(arr) == 1 else 'ies'} stored).")


def cmd_cd_remove(args, wb):
    sheet, row, ws, headers, arr = _load_deadlines(args, wb)
    _check_index(arr, args.index, args.file_no)
    removed = arr.pop(args.index)
    return _finish_deadlines(
        args, ws, row, headers, arr, sheet,
        f"removed deadline at index {args.index}: {json.dumps(removed, ensure_ascii=False)}")


def cmd_cd_resolve(args, wb):
    entry_date = require_iso_date(args.date.strip(), "--date")
    sheet, row, ws, headers, arr = _load_deadlines(args, wb)
    _check_index(arr, args.index, args.file_no)
    entry = arr[args.index]
    if not isinstance(entry, dict) or entry.get("date") or "anchor" not in entry:
        raise contract_error(
            f"Entry at index {args.index} is not an anchored deadline "
            f"({json.dumps(entry, ensure_ascii=False)}) — resolve only converts "
            "anchored entries to dated ones. Use remove + add to change a dated entry.")
    arr[args.index] = {
        "date": entry_date,
        "description": entry.get("description", ""),
        "source": entry.get("source", ""),
    }
    return _finish_deadlines(
        args, ws, row, headers, arr, sheet,
        f"resolved anchored deadline ({entry.get('anchor')!r}) at index "
        f"{args.index} to {entry_date}")


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def emit_success(args, action, file_no, changes, warnings, backup_path,
                 validator, summary):
    for w in warnings:
        print(f"WARNING: {w}", file=sys.stderr)
    if args.json:
        print(json.dumps({
            "ok": True,
            "action": action,
            "file_no": file_no,
            "dry_run": bool(args.dry_run),
            "changes": changes,
            "warnings": warnings,
            "backup": backup_path,
            "validator": validator,
            "summary": summary,
        }, ensure_ascii=False, default=str))
    elif args.dry_run:
        print(f"DRY-RUN {action} {file_no}: would {summary} (no write performed)")
    else:
        print(f"OK {action} {file_no}: {summary} | validator: {validator} "
              f"| backup: {backup_path}")


def run(args, action, mutator):
    tracker = os.path.abspath(args.tracker)
    if not args.dry_run:
        check_lock(tracker, args.force_unlocked)
    wb = open_tracker(tracker)
    try:
        file_no, changes, warnings, summary = mutator(args, wb)
        if args.dry_run:
            # Mutations exist only in memory; the file on disk is untouched.
            emit_success(args, action, file_no, changes, warnings,
                         None, None, summary)
            return EXIT_OK
        backup_path = backup_spreadsheet(tracker)
        atomic_save(wb, tracker)
    finally:
        wb.close()
    validator = run_validator(tracker, backup_path)
    emit_success(args, action, file_no, changes, warnings, backup_path,
                 validator, summary)
    return EXIT_OK


def build_parser():
    common = argparse.ArgumentParser(add_help=False)
    common.add_argument("--tracker", required=True, metavar="PATH",
                        help="Path to the matter tracker .xlsx (required; no default)")
    common.add_argument("--dry-run", action="store_true",
                        help="Report what would change; write nothing")
    common.add_argument("--json", action="store_true",
                        help="Machine-readable JSON result on stdout")
    common.add_argument("--force-unlocked", action="store_true",
                        help="Proceed even if an Excel owner lock file (~$...) exists")

    parser = argparse.ArgumentParser(
        prog="tracker_write.py",
        description="Deterministic write-guard CLI for the matter tracker.")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("update", parents=[common],
                       help="Update cells on the row matching a File #")
    p.add_argument("--file-no", required=True)
    p.add_argument("--set", dest="sets", action="append", required=True,
                   metavar="COLUMN=value",
                   help="Column assignment; repeatable")
    p.set_defaults(func=cmd_update, action="update")

    p = sub.add_parser("timeline", parents=[common],
                       help="Append a dated entry to the Timeline")
    p.add_argument("--file-no", required=True)
    p.add_argument("--date", required=True, metavar="YYYY-MM-DD")
    p.add_argument("--text", required=True)
    p.set_defaults(func=cmd_timeline, action="timeline")

    p = sub.add_parser("new-matter", parents=[common],
                       help="Create a new matter on Open Matters with the next File #")
    p.add_argument("--client", required=True)
    p.add_argument("--description", required=True)
    p.add_argument("--opposing")
    p.add_argument("--email")
    p.add_argument("--phone")
    p.add_argument("--matter-type")
    p.add_argument("--matter-folder")
    p.add_argument("--date-opened", metavar="YYYY-MM-DD",
                   help="Override Date Opened (default: today)")
    p.set_defaults(func=cmd_new_matter, action="new-matter")

    p = sub.add_parser("close", parents=[common],
                       help="Move a matter to Closed Matters")
    p.add_argument("--file-no", required=True)
    p.set_defaults(func=cmd_close, action="close")

    p = sub.add_parser("reopen", parents=[common],
                       help="Move a matter back to Open Matters")
    p.add_argument("--file-no", required=True)
    p.set_defaults(func=cmd_reopen, action="reopen")

    cd = sub.add_parser("court-deadline",
                        help="Manage the Court Deadlines JSON array (col S)")
    cdsub = cd.add_subparsers(dest="cd_command", required=True)

    p = cdsub.add_parser("add", parents=[common],
                         help="Append a dated or anchored deadline")
    p.add_argument("--file-no", required=True)
    p.add_argument("--date", metavar="YYYY-MM-DD")
    p.add_argument("--anchor", metavar="TEXT",
                   help='Anchor event, e.g. "trial date"')
    p.add_argument("--offset-days", type=int, metavar="N",
                   help="Days relative to the anchor (may be negative)")
    p.add_argument("--description", required=True)
    p.add_argument("--source", default="")
    p.set_defaults(func=cmd_cd_add, action="court-deadline add")

    p = cdsub.add_parser("remove", parents=[common],
                         help="Remove a deadline by 0-based index")
    p.add_argument("--file-no", required=True)
    p.add_argument("--index", type=int, required=True)
    p.set_defaults(func=cmd_cd_remove, action="court-deadline remove")

    p = cdsub.add_parser("resolve", parents=[common],
                         help="Convert an anchored deadline to a dated one")
    p.add_argument("--file-no", required=True)
    p.add_argument("--index", type=int, required=True)
    p.add_argument("--date", required=True, metavar="YYYY-MM-DD")
    p.set_defaults(func=cmd_cd_resolve, action="court-deadline resolve")

    return parser


def main(argv=None):
    args = build_parser().parse_args(argv)
    try:
        return run(args, args.action, args.func)
    except GuardError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        if getattr(args, "json", False):
            print(json.dumps({"ok": False, "error": str(e),
                              "exit_code": e.code}, ensure_ascii=False))
        return e.code


if __name__ == "__main__":
    sys.exit(main())
